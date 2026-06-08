"""
MÓDULO DE EXPORTAÇÃO - Planilha Excel (.xlsx)
Gera planilha editável com todos os dados e cálculos do projeto PPCI.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


AZUL_ESCURO = "1B3A5C"
AZUL_CLARO = "D6E4F0"
LARANJA = "F4A460"
BRANCO = "FFFFFF"
CINZA_CLARO = "F2F2F2"

FONT_TITULO = Font(name="Arial", size=14, bold=True, color=BRANCO)
FONT_SUBTITULO = Font(name="Arial", size=11, bold=True, color=AZUL_ESCURO)
FONT_HEADER = Font(name="Arial", size=10, bold=True, color=BRANCO)
FONT_NORMAL = Font(name="Arial", size=10)
FONT_RESULTADO = Font(name="Arial", size=10, bold=True, color="D84315")
FILL_TITULO = PatternFill("solid", fgColor=AZUL_ESCURO)
FILL_HEADER = PatternFill("solid", fgColor="2E5A88")
FILL_ALTERNADA = PatternFill("solid", fgColor=CINZA_CLARO)
FILL_RESULTADO = PatternFill("solid", fgColor="FFF3E0")
BORDER_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _add_titulo(ws, row, text, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_TITULO
    cell.fill = FILL_TITULO
    cell.alignment = ALIGN_CENTER
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).fill = FILL_TITULO
        ws.cell(row=row, column=c).border = BORDER_THIN


def _add_header_row(ws, row, headers):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN


def _add_data_row(ws, row, values, alt=False):
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = FONT_NORMAL
        cell.alignment = ALIGN_LEFT
        cell.border = BORDER_THIN
        if alt:
            cell.fill = FILL_ALTERNADA


def gerar_planilha(projeto: dict, caminho: str) -> str:
    """
    Gera planilha Excel completa do projeto PPCI.
    
    projeto: dict com resultado do endpoint /api/calculos/projeto-completo
    caminho: path do arquivo de saída (.xlsx)
    """
    wb = Workbook()

    # ============================================================
    # ABA 1: DADOS GERAIS
    # ============================================================
    ws = wb.active
    ws.title = "Dados Gerais"
    _set_col_widths(ws, {"A": 30, "B": 35, "C": 20, "D": 20, "E": 20, "F": 20})

    _add_titulo(ws, 1, "PROJETO PPCI — DADOS GERAIS")
    ws.row_dimensions[1].height = 30

    r = 3
    cls = projeto.get("classificacao", {})
    ocup = cls.get("ocupacao", {})
    risco = cls.get("risco", {})
    proc = cls.get("processo", {})

    dados_gerais = [
        ("Nome do Projeto", projeto.get("nome_projeto", "")),
        ("Data de Geração", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("", ""),
        ("CLASSIFICAÇÃO", ""),
        ("Grupo de Ocupação", f"{ocup.get('grupo', '')} — {ocup.get('nome_grupo', '')}"),
        ("Divisão", f"{ocup.get('divisao', '')} — {ocup.get('descricao', '')}"),
        ("Nível de Risco", risco.get("descricao", "")),
        ("Carga de Incêndio Adotada", f"{cls.get('carga_incendio', {}).get('adotada', 0)} MJ/m²"),
        ("População Adotada", f"{cls.get('populacao', {}).get('adotada', 0)} pessoas"),
        ("Tipo de Processo", f"{proc.get('tipo', '')} — {proc.get('descricao', '')}"),
        ("", ""),
        ("SISTEMAS EXIGIDOS", f"{cls.get('total_sistemas', 0)} sistemas"),
    ]

    for item in dados_gerais:
        if item[0] in ("CLASSIFICAÇÃO", "SISTEMAS EXIGIDOS"):
            ws.cell(row=r, column=1, value=item[0]).font = FONT_SUBTITULO
            ws.cell(row=r, column=2, value=item[1]).font = FONT_RESULTADO
        else:
            ws.cell(row=r, column=1, value=item[0]).font = Font(name="Arial", size=10, bold=True)
            ws.cell(row=r, column=2, value=item[1]).font = FONT_NORMAL
        r += 1

    r += 1
    sistemas = cls.get("sistemas_exigidos", [])
    if sistemas:
        _add_header_row(ws, r, ["Código", "Sistema", "IT Referência", "", "", ""])
        r += 1
        for i, s in enumerate(sistemas):
            _add_data_row(ws, r, [s.get("codigo", ""), s.get("nome", ""), s.get("it_referencia", ""), "", "", ""], alt=i % 2 == 0)
            r += 1

    # ============================================================
    # ABA 2: SAÍDA DE EMERGÊNCIA
    # ============================================================
    ws2 = wb.create_sheet("Saída de Emergência")
    _set_col_widths(ws2, {"A": 35, "B": 50, "C": 15, "D": 15, "E": 20, "F": 20})

    _add_titulo(ws2, 1, "DIMENSIONAMENTO — SAÍDAS DE EMERGÊNCIA (IT-11)")
    ws2.row_dimensions[1].height = 30

    saida = projeto.get("saida_emergencia", {})
    etapas = saida.get("etapas", [])

    _add_header_row(ws2, 3, ["Etapa", "Fórmula / Critério", "Valor", "Unidade", "Justificativa", ""])
    r = 4
    for i, e in enumerate(etapas):
        _add_data_row(ws2, r, [
            e.get("etapa", ""),
            e.get("formula", ""),
            str(e.get("valor", "")),
            e.get("unidade", ""),
            e.get("justificativa", ""),
            "",
        ], alt=i % 2 == 0)
        r += 1

    resumo_saida = saida.get("resumo", {})
    if resumo_saida:
        r += 1
        ws2.cell(row=r, column=1, value="RESUMO").font = FONT_SUBTITULO
        r += 1
        for k, v in resumo_saida.items():
            ws2.cell(row=r, column=1, value=k.replace("_", " ").title()).font = FONT_NORMAL
            val = str(v) if not isinstance(v, dict) else str(v.get("descricao", v))
            ws2.cell(row=r, column=2, value=val).font = FONT_RESULTADO
            r += 1

    # ============================================================
    # ABA 3: EXTINTORES
    # ============================================================
    ws3 = wb.create_sheet("Extintores")
    _set_col_widths(ws3, {"A": 35, "B": 50, "C": 15, "D": 15, "E": 20, "F": 20})
    _add_titulo(ws3, 1, "DIMENSIONAMENTO — EXTINTORES (IT-04)")

    ext = projeto.get("extintores", {})
    etapas_ext = ext.get("etapas", [])
    _add_header_row(ws3, 3, ["Etapa", "Fórmula / Critério", "Valor", "Unidade", "Justificativa", ""])
    r = 4
    for i, e in enumerate(etapas_ext):
        val = e.get("valor", "")
        if isinstance(val, list):
            val = ", ".join(str(x) for x in val)
        _add_data_row(ws3, r, [e.get("etapa", ""), e.get("formula", ""), str(val), e.get("unidade", ""), e.get("justificativa", ""), ""], alt=i % 2 == 0)
        r += 1

    # ============================================================
    # ABA 4: HIDRANTES
    # ============================================================
    hid = projeto.get("hidrantes")
    if hid:
        ws4 = wb.create_sheet("Hidrantes")
        _set_col_widths(ws4, {"A": 35, "B": 50, "C": 15, "D": 15, "E": 20, "F": 20})
        _add_titulo(ws4, 1, "DIMENSIONAMENTO — HIDRANTES (IT-17)")
        _add_header_row(ws4, 3, ["Etapa", "Fórmula / Critério", "Valor", "Unidade", "Justificativa", ""])
        r = 4
        for i, e in enumerate(hid.get("etapas", [])):
            _add_data_row(ws4, r, [e.get("etapa", ""), e.get("formula", ""), str(e.get("valor", "")), e.get("unidade", ""), e.get("justificativa", ""), ""], alt=i % 2 == 0)
            r += 1

    # ============================================================
    # ABA 5: RESERVA + BOMBA
    # ============================================================
    res = projeto.get("reserva_tecnica")
    bom = projeto.get("bomba_incendio")
    if res or bom:
        ws5 = wb.create_sheet("Reserva e Bomba")
        _set_col_widths(ws5, {"A": 35, "B": 50, "C": 15, "D": 15, "E": 20, "F": 20})
        _add_titulo(ws5, 1, "RESERVA TÉCNICA E BOMBA DE INCÊNDIO")
        r = 3
        if res:
            ws5.cell(row=r, column=1, value="RESERVA TÉCNICA DE INCÊNDIO (IT-17/IT-22)").font = FONT_SUBTITULO
            r += 1
            _add_header_row(ws5, r, ["Etapa", "Fórmula / Critério", "Valor", "Unidade", "", ""])
            r += 1
            for i, e in enumerate(res.get("etapas", [])):
                _add_data_row(ws5, r, [e.get("etapa", ""), e.get("formula", ""), str(e.get("valor", "")), e.get("unidade", ""), "", ""], alt=i % 2 == 0)
                r += 1
            r += 1

        if bom:
            ws5.cell(row=r, column=1, value="BOMBA DE INCÊNDIO (IT-22)").font = FONT_SUBTITULO
            r += 1
            _add_header_row(ws5, r, ["Etapa", "Fórmula / Critério", "Valor", "Unidade", "", ""])
            r += 1
            for i, e in enumerate(bom.get("etapas", [])):
                _add_data_row(ws5, r, [e.get("etapa", ""), e.get("formula", ""), str(e.get("valor", "")), e.get("unidade", ""), "", ""], alt=i % 2 == 0)
                r += 1

    wb.save(caminho)
    return caminho
